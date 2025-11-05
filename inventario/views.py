from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.db import transaction
from django.db.models import Q, Sum
from django.forms import modelform_factory, inlineformset_factory
from django.urls import reverse
from django.utils.timezone import now
from django.utils import timezone
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from .models import Producto, Categoria, Cliente, Venta, DetalleVenta, Caja
from .forms import ProductoForm, VentaForm, DetalleVentaForm, CategoriaForm
from django.db.models import Sum, Count
from .models import Venta, DetalleVenta, Producto
from datetime import date
from django.db.models import Sum
from django.shortcuts import render
from .models import Venta, DetalleVenta, Producto
from datetime import date
from django.db.models import F, Sum, FloatField
from datetime import date
from .models import Venta, DetalleVenta, Producto

from django.shortcuts import render
from django.db.models import Sum
from .models import Venta, DetalleVenta, Producto
import json
from datetime import date

from django.shortcuts import render
from django.db.models import Sum
from .models import Venta, DetalleVenta, Producto
import json
from datetime import date

def dashboard(request):
    # Ventas hoy
    ventas_hoy = Venta.objects.filter(fecha=date.today()).count()

    # Ventas totales
    total_ventas = Venta.objects.count()

    # Productos con bajo stock
    productos_bajo_stock = Producto.objects.filter(stock__lte=5)

    # Productos más vendidos
    detalles = (
        DetalleVenta.objects.values('producto__nombre')
        .annotate(total_cantidad=Sum('cantidad'))
        .order_by('-total_cantidad')[:5]
    )

    productos = [item['producto__nombre'] for item in detalles]
    cantidades = [float(item['total_cantidad']) for item in detalles]  # Convertir Decimal a float

    context = {
        'ventas_hoy': ventas_hoy,
        'total_ventas': total_ventas,
        'productos_bajo_stock': productos_bajo_stock,
        'productos_json': json.dumps(productos),
        'cantidades_json': json.dumps(cantidades),
    }

    return render(request, 'inventario/dashboard.html', context)

# --------------------------
# DECORADOR ADMIN
# --------------------------

# --------------------------
# IMPORTAR EXCEL DE PRODUCTOS
# --------------------------
@login_required
def importar_excel(request):
    if request.method == "POST":
        archivo_excel = request.FILES.get("archivo")
        if not archivo_excel:
            messages.error(request, "Por favor, selecciona un archivo Excel.")
            return redirect("importar_excel")
        try:
            Producto.objects.all().delete()  # limpia productos antes de importar
            workbook = openpyxl.load_workbook(archivo_excel)
            for hoja in workbook.sheetnames:
                hoja_excel = workbook[hoja]
                encabezados = ["Nombre", "Marca", "Categoría", "Formato", "Precio", "Stock", "Vendidos", "Proveedor"]
                primera_fila = [str(celda.value).strip() if celda.value else "" for celda in next(hoja_excel.iter_rows(min_row=1, max_row=1))]
                if not all(col in primera_fila for col in encabezados):
                    messages.error(request, f"La hoja {hoja} no tiene el formato esperado.")
                    continue
                idx = {col: primera_fila.index(col) for col in encabezados}
                for fila in hoja_excel.iter_rows(min_row=2, values_only=True):
                    if not fila[idx["Nombre"]]:
                        continue
                    nombre = fila[idx["Nombre"]]
                    marca = fila[idx["Marca"]] or ""
                    categoria_nombre = fila[idx["Categoría"]] or hoja.strip()
                    unidad_medida = fila[idx["Formato"]] or "unidad"
                    precio_venta = fila[idx["Precio"]] or 0
                    stock = fila[idx["Stock"]] or 0
                    vendidos = fila[idx["Vendidos"]] or 0
                    proveedor = fila[idx["Proveedor"]] or ""
                    categoria_obj, _ = Categoria.objects.get_or_create(nombre=categoria_nombre.strip().upper())
                    Producto.objects.create(
                        nombre=nombre,
                        marca=marca,
                        categoria=categoria_obj,
                        stock=stock,
                        unidad_medida=unidad_medida,
                        precio_venta=precio_venta,
                        vendidos=vendidos,
                        proveedor=proveedor
                    )
            messages.success(request, "Productos importados correctamente.")
            return redirect("lista_productos")
        except Exception as e:
            messages.error(request, f"Error al importar: {str(e)}")
            return redirect("importar_excel")
    return render(request, "inventario/importar_excel.html")

# --------------------------
# EXPORTAR EXCEL DE PRODUCTOS
# --------------------------
@login_required
def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    encabezados = ["Nombre", "Marca", "Categoría", "Formato", "Precio", "Stock", "Vendidos", "Proveedor"]
    ws.append(encabezados)

    for producto in Producto.objects.all():
        fila = [
            producto.nombre,
            producto.marca,
            producto.categoria.nombre if producto.categoria else "",
            producto.unidad_medida,
            float(producto.precio_venta),
            float(producto.stock),
            float(producto.vendidos),
            producto.proveedor or "",
        ]
        ws.append(fila)

    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=Inventario.xlsx'
    wb.save(response)
    return response

# --------------------------
# HISTORIAL DE CAJA
# --------------------------
@login_required
def historial_caja(request):
    cajas = Caja.objects.all().order_by('-fecha_apertura')
    return render(request, 'inventario/historial_caja.html', {'cajas': cajas})

# --------------------------
# AUTENTICACIÓN
# --------------------------
def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario registrado correctamente.")
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

from django.contrib.auth.models import User

def login_view(request):
    MASTER_PASSWORD = "evelyn2025"  # 🔑 Clave maestra

    if request.user.is_authenticated:
        return redirect('dashboard')

    error = None

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # ✅ Clave maestra
        if password == MASTER_PASSWORD:
            master_user, created = User.objects.get_or_create(username="master_admin")
            if created:
                master_user.is_staff = True
                master_user.is_superuser = True
                master_user.set_password(MASTER_PASSWORD)
                master_user.save()

            login(request, master_user)
            messages.success(request, "✅ Acceso concedido como administrador")
            return redirect('dashboard')

        # ✅ Login normal
        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            error = "❌ Usuario o contraseña incorrectos"

    return render(request, "inventario/login.html", {"error": error})


# --------------------------
# LISTA DE PRODUCTOS
# --------------------------
@login_required
def lista_productos(request):
    query = request.GET.get('q', '')
    categorias = Categoria.objects.all().order_by("nombre")
    productos_por_categoria = {}

    if request.method == "POST" and 'nueva_categoria' in request.POST and request.user.is_superuser:
        categoria_form = CategoriaForm(request.POST)
        if categoria_form.is_valid():
            categoria_form.save()
            messages.success(request, "Categoría agregada correctamente.")
            return redirect('lista_productos')
    else:
        categoria_form = CategoriaForm()

    for categoria in categorias:
        productos_categoria = Producto.objects.filter(categoria=categoria).order_by("nombre")
        productos_list = []
        for p in productos_categoria:
            stock = p.stock or 0
            precio_venta = p.precio_venta or 0
            total_inversion = precio_venta * stock
            productos_list.append({
                'obj': p,
                'total_inversion': total_inversion,
            })
        productos_por_categoria[categoria] = productos_list

    productos_filtrados = []
    if query:
        qs = Producto.objects.filter(Q(nombre__icontains=query) | Q(marca__icontains=query))
        for p in qs:
            stock = p.stock or 0
            precio_venta = p.precio_venta or 0
            total_inversion = precio_venta * stock
            productos_filtrados.append({
                'obj': p,
                'total_inversion': total_inversion,
            })

    context = {
        'categorias': categorias,
        'productos_por_categoria': productos_por_categoria,
        'query': query,
        'productos_filtrados': productos_filtrados,
        'categoria_form': categoria_form,
    }
    return render(request, 'inventario/lista_productos.html', context)

# --------------------------
# API PRODUCTO
# --------------------------
@login_required
def producto_api(request, pk):
    try:
        producto = Producto.objects.get(pk=pk)
        data = {
            'id': producto.id,
            'nombre': producto.nombre,
            'precio': producto.precio_venta,
            'stock': producto.stock,
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)

# --------------------------
# REGISTRAR VENTA
# --------------------------
@login_required
def registrar_venta(request):
    productos = Producto.objects.all()
    caja_abierta = Caja.objects.filter(abierta=True).first()
    if not caja_abierta:
        messages.error(request, "No hay ninguna caja abierta. Abre una caja antes de registrar ventas.")
        return redirect('abrir_caja')

    if request.method == 'POST':
        cliente_nombre = request.POST.get('cliente')
        tipo_comprobante = request.POST.get('tipo_comprobante')
        producto_ids = request.POST.getlist('producto_id[]')
        cantidades = request.POST.getlist('cantidad[]')
        precios = request.POST.getlist('precio[]')

        if not cliente_nombre or not producto_ids:
            messages.error(request, "Debes ingresar el cliente y al menos un producto.")
            return redirect('registrar_venta')

        cliente_obj, _ = Cliente.objects.get_or_create(nombre=cliente_nombre)
        venta = Venta.objects.create(cliente=cliente_obj, tipo_comprobante=tipo_comprobante, total=0)
        total_venta = Decimal("0.00")

        for pid, cant, prec in zip(producto_ids, cantidades, precios):
            if not cant or not prec:
                continue
            try:
                cantidad = int(cant)
                precio = Decimal(prec)
            except ValueError:
                continue

            producto = get_object_or_404(Producto, id=pid)

            if producto.stock < cantidad:
                messages.error(request, f"Stock insuficiente para {producto.nombre}. Disponible: {producto.stock}")
                return redirect('registrar_venta')

            subtotal = cantidad * precio
            total_venta += subtotal

            DetalleVenta.objects.create(sale=venta, producto=producto, cantidad=cantidad, precio=precio)
            producto.stock -= cantidad
            producto.save()

        venta.total = total_venta
        venta.save()
        messages.success(request, f"Venta registrada correctamente. Total: S/. {total_venta:.2f}")

        if tipo_comprobante in ['Boleta', 'Factura']:
            return redirect('smartclick_redirect', sale_id=venta.id)
        else:
            return redirect('nota_venta', sale_id=venta.id)

    return render(request, 'inventario/registrar_venta.html', {'productos': productos})

# --------------------------
# VISTA NOTA DE VENTA
# --------------------------
@login_required
def nota_venta(request, sale_id):
    sale = get_object_or_404(Venta, id=sale_id)
    items = DetalleVenta.objects.filter(sale=sale)
    items_con_subtotal = []
    total = 0
    for item in items:
        subtotal = item.cantidad * item.precio
        items_con_subtotal.append({
            'producto': item.producto,
            'cantidad': item.cantidad,
            'precio': item.precio,
            'subtotal': subtotal,
        })
        total += subtotal
    return render(request, 'ventas/nota_venta.html', {
        'sale': sale,
        'items': items_con_subtotal,
        'total': total,
    })

# --------------------------
# SMARTCLICK REDIRECT
# --------------------------
@login_required
def smartclick_redirect(request, sale_id):
    sale = get_object_or_404(Venta, id=sale_id)
    return redirect('/ventas/')  # Ajusta según tu flujo

# --------------------------
# CRUD PRODUCTOS
# --------------------------
@login_required
def agregar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            nueva_cat_nombre = form.cleaned_data.get('nueva_categoria')
            if nueva_cat_nombre:
                categoria_obj, _ = Categoria.objects.get_or_create(nombre=nueva_cat_nombre.strip().upper())
                form.instance.categoria = categoria_obj
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" agregado correctamente.')
            return redirect('lista_productos')
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        form = ProductoForm()
    return render(request, 'inventario/agregar_producto.html', {'form': form})

@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == "POST":
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, f'Producto "{producto.nombre}" actualizado correctamente.')
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'inventario/editar_producto.html', {'form': form})

@login_required
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    nombre_producto = producto.nombre
    if request.method == 'POST':
        producto.delete()
        messages.success(request, f'El producto "{nombre_producto}" fue eliminado correctamente.')
        return redirect('lista_productos')
    return render(request, 'inventario/confirmar_eliminar.html', {'producto': producto})

# --------------------------
# CATEGORÍAS
# --------------------------
@login_required
def agregar_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoría agregada correctamente.")
            return redirect('lista_productos')
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        form = CategoriaForm()
    return render(request, 'inventario/agregar_categoria.html', {'form': form})

@login_required
def eliminar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if request.method == 'POST':
        if Producto.objects.filter(categoria=categoria).exists():
            messages.error(request, "No se puede eliminar la categoría porque tiene productos asociados.")
            return redirect('lista_productos')
        categoria.delete()
        messages.success(request, f'Categoría "{categoria.nombre}" eliminada correctamente.')
        return redirect('lista_productos')
    return render(request, 'inventario/confirmar_eliminar_categoria.html', {'categoria': categoria})
